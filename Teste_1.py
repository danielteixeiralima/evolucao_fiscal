# -*- coding: utf-8 -*-
import re
import sys
import os
import json
import time
from time import perf_counter
from time import time as _now
from datetime import datetime
from typing import List, Tuple, Dict, Any, Optional
from concurrent.futures import ThreadPoolExecutor, as_completed
from threading import local, Lock

import requests
from requests.auth import HTTPBasicAuth
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry

import pandas as pd
from pandas import json_normalize
from dateutil import tz, parser as dtparser
import argparse
import tempfile, shutil
# ==============================
# CONFIG DE CONEXÃO
# ==============================
HOST = "http://192.168.18.9:8051"
USER = "INTEGRA_INOVAI"
PWD  = "INOVAI.LAB"
auth = HTTPBasicAuth(USER, PWD)

MOV_ENDPOINT = "/api/mov/v1/movements"

# ==============================
# AJUSTES DE REDE / TOLERÂNCIA
# ==============================
MAX_LIST_PAGES           = 3
MAX_PAGES_PER_BRANCH     = 1_000_000

CONNECT_TIMEOUT          = 10
READ_TIMEOUT             = 45
ADAPTER_RETRIES          = 3
MANUAL_RETRIES           = 2
BACKOFF_BASE             = 0.5

BRANCH_DEADLINE_SEC_DEFAULT = 0  # 0 = sem limite (não perde itens)

DEBUG_VERBOSE            = True

# Campos típicos
DATE_FIELD_HINTS = [
    "date","issueDate","emissionDate","movementDate","creationDate","createdDate",
    "entryDate","exitDate","lastEditTime","registrationDate",
    "Date","IssueDate","EmissionDate","MovementDate","CreationDate","CreatedDate",
    "EntryDate","ExitDate","LastEditTime","RegistrationDate"
]
COMPANY_KEYS = ["companyId","CompanyId","companyCode","CompanyCode"]
BRANCH_KEYS  = ["branchId","BranchId","branchCode","BranchCode"]

# $select padrão (pode desligar)
DEFAULT_SELECT_FIELDS = [
    "date","internalId","movementId","documentNumber","type","grossValue","netValue",
    "status","companyId","branchId","partner.name"
]

# ==============================
# HTTP por thread (keep-alive + gzip)
# ==============================
_thread_local = local()

def _build_session() -> requests.Session:
    retry = Retry(
        total=ADAPTER_RETRIES, connect=ADAPTER_RETRIES, read=ADAPTER_RETRIES,
        backoff_factor=0.4, status_forcelist=[502,503,504],
        allowed_methods=["GET","HEAD"], raise_on_status=False,
    )
    adapter = HTTPAdapter(max_retries=retry, pool_connections=64, pool_maxsize=256)
    s = requests.Session()
    s.auth = auth
    s.headers.update({
        "Accept": "application/json",
        "Accept-Encoding": "gzip, deflate",
        # não forçar Connection: close → mantém keep-alive
    })
    s.mount("http://", adapter); s.mount("https://", adapter)
    return s

def _sess() -> requests.Session:
    s = getattr(_thread_local, "session", None)
    if s is None:
        s = _build_session()
        _thread_local.session = s
    return s

def _safe_request(method: str, url: str, *, params=None, timeout=None):
    print(f"DEBUG: chamando {method} {url} com params={params}")

    timeout = timeout or (CONNECT_TIMEOUT, READ_TIMEOUT)
    last_exc = None
    for attempt in range(1, MANUAL_RETRIES + 1):
        t0 = perf_counter()
        try:
            r = _sess().request(method, url, params=params, timeout=timeout)
            elapsed = (perf_counter() - t0) * 1000
            if DEBUG_VERBOSE:
                print(f"      ↪ {method} {url}  ({elapsed:.0f} ms)")
                if params:
                    print(f"        params: {params}")
                print(f"        → HTTP {r.status_code}")
            return r, None, elapsed
        except (requests.exceptions.ReadTimeout, requests.exceptions.ConnectionError) as e:
            elapsed = (perf_counter() - t0) * 1000
            last_exc = e
            sleep_s = BACKOFF_BASE * (2 ** (attempt - 1))
            print(f"      ⚠️ {type(e).__name__} {attempt}/{MANUAL_RETRIES} ({elapsed:.0f} ms). Backoff {sleep_s:.1f}s…")
            time.sleep(sleep_s)
        except Exception as e:
            elapsed = (perf_counter() - t0) * 1000
            print(f"      ❌ Exception inesperada ({elapsed:.0f} ms): {type(e).__name__}: {e}")
            return None, e, elapsed
    return None, last_exc, 0.0

# ==============================
# UTILS
# ==============================
def _norm_code(x: str) -> str:
    s = str(x).strip()
    return str(int(s)) if s.isdigit() else s

def _detect_field_present(obj: Dict[str,Any], candidates: List[str]) -> Optional[str]:
    lower_map = {k.lower(): k for k in obj.keys()}
    for cand in candidates:
        if cand.lower() in lower_map:
            return lower_map[cand.lower()]
    for k in obj.keys():
        if any(re.search(fr"\b{re.escape(cand)}\b", k, flags=re.IGNORECASE) for cand in candidates):
            return k
    return None

def _build_company_branch_filter(company_field: Optional[str], branch_field: Optional[str], comp_code: str, branch_code: str) -> Optional[str]:
    parts=[]
    if company_field:
        val = _norm_code(comp_code)
        parts.append(f"{company_field} eq {val if val.isdigit() else repr(val)}")
    if branch_field and str(branch_code).strip() != "":
        bc = str(branch_code).strip()
        parts.append(f"{branch_field} eq {bc if bc.isdigit() else repr(bc)}")
    return " and ".join(parts) if parts else None

def _date_iso(dt: datetime) -> str:
    return dt.isoformat()

def _guess_datetime_fields_from_sample(items: List[Dict[str,Any]]) -> List[str]:
    found=set()
    for it in items:
        for k,v in it.items():
            if isinstance(v,str):
                try:
                    dtparser.isoparse(v); found.add(k)
                except Exception:
                    pass
    ordered = [f for f in DATE_FIELD_HINTS if f in found]
    for f in sorted(found):
        if f not in ordered:
            ordered.append(f)
    return ordered

def _coerce_datetime_cols(df: pd.DataFrame) -> pd.DataFrame:
    candidates = [
        "date","Date","movementDate","MovementDate","issueDate","IssueDate",
        "emissionDate","EmissionDate","creationDate","CreatedDate","entryDate","exitDate","lastEditTime"
    ]
    for c in candidates:
        if c in df.columns:
            df[c] = pd.to_datetime(df[c], errors="coerce", utc=False)
    return df

def _strip_timezone_for_excel(df: pd.DataFrame, target_tz: str = "America/Sao_Paulo") -> pd.DataFrame:
    from pandas.api.types import is_datetime64_any_dtype
    for col in df.columns:
        if is_datetime64_any_dtype(df[col]):
            tzinfo = getattr(df[col].dtype, "tz", None)
            if tzinfo is not None:
                try:
                    df[col] = df[col].dt.tz_convert(target_tz).dt.tz_localize(None)
                except Exception:
                    df[col] = df[col].dt.tz_localize(None)
    return df

def _jsonify_non_scalars_for_excel(df: pd.DataFrame) -> pd.DataFrame:
    def to_excel_scalar(x):
        if isinstance(x, (list, dict, tuple, set)):
            try:
                return json.dumps(x, ensure_ascii=False)
            except Exception:
                return str(x)
        return x
    for c in df.columns:
        df[c] = df[c].apply(to_excel_scalar)
    return df

def _first_existing(cols: list[str], df_cols: set[str]) -> Optional[str]:
    for c in cols:
        if c in df_cols: return c
    return None

def _smart_sort(df: pd.DataFrame, ordenar_cols: List[str]) -> pd.DataFrame:
    cols = [c for c in ordenar_cols if c in df.columns]
    if not cols:
        maybe_date = _first_existing(["date","Date","movementDate","MovementDate","issueDate","IssueDate"], set(df.columns))
        maybe_id   = _first_existing(["internalId","InternalId","movementId","MovementId"], set(df.columns))
        cols = [c for c in [maybe_date, maybe_id] if c]
    if cols:
        return df.sort_values(cols, kind="stable")
    return df

def _autosize_openpyxl(ws):
    from openpyxl.utils import get_column_letter
    dims = {}
    for row in ws.rows:
        for cell in row:
            if cell.value is None: continue
            val = str(cell.value)
            dims[cell.column] = max(dims.get(cell.column, 0), len(val))
    for col, width in dims.items():
        ws.column_dimensions[get_column_letter(col)].width = min(max(width + 2, 10), 60)

# ==============================
# CATÁLOGO
# ==============================
def get_company_name(company_code: str) -> str:
    # evita chamadas para /api/framework/v1/companies
    return f"Empresa {company_code}"


def list_branches(company_code: str) -> list[dict]:
    # usa sempre a filial informada por parâmetro
    return [{"CompanyCode": str(company_code), "Code": args.filial or "1", "Title": "Filial simulada"}]


def list_companies() -> List[Dict[str, str]]:
    urls = [
        f"{HOST}/api/framework/v1/companies",
        f"{HOST}/api/framework/v1/Companies",
        f"{HOST}/api/corp/v1/companies",
    ]
    out: List[Dict[str,str]] = []
    seen = set()
    for url in urls:
        r, exc, _ = _safe_request("GET", url)
        if not r or r.status_code != 200:
            continue
        try:
            j = r.json()
        except Exception:
            continue
        items = j.get("items", j) if isinstance(j, dict) else j
        if not isinstance(items, list):
            continue
        for obj in items:
            code = str(obj.get("code") or obj.get("Code") or obj.get("companyCode") or obj.get("CompanyCode") or "").strip()
            name = str(obj.get("name") or obj.get("Name") or obj.get("title") or obj.get("Title") or obj.get("description") or "").strip()
            if not code: continue
            code = _norm_code(code)
            if code in seen: continue
            seen.add(code)
            out.append({"Code": code, "Name": name})
        if out:
            break
    return out



# ==============================
# MOVEMENTS
# ==============================


def get_customer_vendor_info(code: str, company_id: str) -> Optional[Dict[str, Any]]:
    if not code:
        return None

    url = f"{HOST}/api/fin/v1/CustomerVendor"
    params = {
        "$filter": f"code eq '{code}' and companyId eq {company_id}",
        "$select": "code,shortName,name,mainNIF"
    }

    r, exc, _ = _safe_request("GET", url, params=params)
    if not r or r.status_code != 200:
        return None

    j = r.json()
    items = j.get("items", [])
    if items:
        return items[0]

    # 🔄 fallback: tenta sem companyId
    params["$filter"] = f"code eq '{code}'"
    r, exc, _ = _safe_request("GET", url, params=params)
    if not r or r.status_code != 200:
        return None

    j = r.json()
    return (j.get("items") or [None])[0]


def _get_movements_page(page: int, page_size: int, odata_filter: Optional[str], select_fields: List[str]):
    """
    Executa a chamada à API /movements, garantindo que o filtro de empresa/filial e de movimentos
    seja combinado corretamente sem ser sobrescrito.
    """
    url = f"{HOST}{MOV_ENDPOINT}"
    params = {
        "page": max(1, page),
        "pageSize": max(1, page_size)
    }

    # Filtro base (empresa + filial)
    base_filter = odata_filter or ""

    # Filtro adicional de movimento
    if getattr(args, "movimento", None):
        movimentos = [m.strip() for m in args.movimento.split(",") if m.strip()]
        if len(movimentos) == 1:
            mov_filtro = f"movementTypeCode eq '{movimentos[0]}'"
        else:
            clauses = [f"movementTypeCode eq '{m}'" for m in movimentos]
            mov_filtro = " or ".join(clauses)

        # ✅ Garante que o filtro de movimento é somado, não sobrescrito
        if base_filter:
            base_filter = f"({base_filter}) and ({mov_filtro})"
        else:
            base_filter = mov_filtro

    # Aplica filtro final
    if base_filter:
        params["$filter"] = base_filter

    # Campos selecionados
    if select_fields:
        params["$select"] = ",".join(select_fields)

    if DEBUG_VERBOSE:
        print(f"DEBUG: chamando GET {url}")
        print(f"       params = {json.dumps(params, indent=2, ensure_ascii=False)}")

    # Execução HTTP
    r, exc, _ = _safe_request("GET", url, params=params)
    meta = {"status": r.status_code if r else -1}

    if not r:
        return [], -1, meta

    if 200 <= r.status_code < 300 or r.status_code == 206:
        try:
            j = r.json()
        except Exception:
            return [], r.status_code, meta

        # Normaliza retorno
        if isinstance(j, dict):
            data = j.get("items") or j.get("Items")
            items = data if isinstance(data, list) else ([j] if not isinstance(j, list) else [])
        elif isinstance(j, list):
            items = j
        else:
            items = []

        return items, r.status_code, meta

    return [], r.status_code, meta


def _try_date_filter_server_side(company_field: Optional[str], branch_field: Optional[str],
                                 comp_code: str, branch_code: str,
                                 dt_ini: datetime, dt_fim: datetime,
                                 sample_date_candidates: List[str],
                                 select_fields: List[str]) -> Tuple[Optional[str], Optional[str]]:
    base_filter = _build_company_branch_filter(company_field, branch_field, comp_code, branch_code)
    candidates = list(dict.fromkeys(DATE_FIELD_HINTS + sample_date_candidates))
    for df in candidates:
        date_part = f"{df} ge {_date_iso(dt_ini)} and {df} le {_date_iso(dt_fim)}"
        odata = f"{base_filter} and {date_part}" if base_filter else date_part
        items, st, _ = _get_movements_page(1, 1, odata, select_fields)
        if st == 200:
            return odata, df
    return None, None

# ==============================
# EXTRAÇÃO SEQUENCIAL POR PÁGINAS (estável e rápida)
# ==============================
def extract_branch_items(comp_norm: str, b_code: str,
                         comp_field: str, br_field: str,
                         dt_ini: datetime, dt_fim: datetime,
                         date_candidates: List[str],
                         page_size_first: int,
                         select_fields: List[str],
                         deadline_sec: int) -> List[Dict[str,Any]]:
    header = f"  → Filial código {b_code or '(vazio)'}"
    print("\n" + header)
    print("  " + "—"*len(header))

    server_odata, used_date_field = _try_date_filter_server_side(
        comp_field, br_field, comp_norm, b_code, dt_ini, dt_fim, date_candidates, select_fields
    )
    if server_odata:
        print(f"    ✅ Filtro server-side aceito: {server_odata}")
    else:
        server_odata = _build_company_branch_filter(comp_field, br_field, comp_norm, b_code)
        print("    ⚠️ Sem campo de data aceito no servidor. Período será filtrado no cliente.")

    # cadeia degradável de page sizes
    ps_chain = [max(1, page_size_first), 50, 25, 10, 5, 1]
    ps_chain = [ps_chain[0]] + [p for p in ps_chain[1:] if p < ps_chain[0]]

    page = 1
    ps_index = 0
    empty_seq = 0
    started_at = _now()
    out: List[Dict[str,Any]] = []

    # dedupe por filial (id pode repetir entre páginas em alguns backends)
    seen_ids = set()
    id_fields = ["internalId","InternalId","movementId","MovementId","id","Id"]

    def in_period_local(rec: Dict[str, Any]) -> bool:
        v = rec.get("date")
        if isinstance(v, str):
            try:
                d = dtparser.isoparse(v)
                return dt_ini <= d <= dt_fim
            except Exception:
                pass
        for key in (date_candidates or []) + DATE_FIELD_HINTS:
            v = rec.get(key)
            if isinstance(v, str):
                try:
                    d = dtparser.isoparse(v)
                    if dt_ini <= d <= dt_fim:
                        return True
                except Exception:
                    pass
        return False
    global_limit = getattr(args, "limit", 0)

    max_pages = getattr(args, "max_pages", 1)
    while page <= min(max_pages, MAX_PAGES_PER_BRANCH):

        if deadline_sec and (_now() - started_at > deadline_sec):
            print(f"      ⏱️ Tempo excedido ({deadline_sec}s) — encerrando filial com {len(out)} itens.")
            break

        page_size = ps_chain[ps_index]
        items, st, _ = _get_movements_page(page, page_size, server_odata, select_fields)

        if st == -1 or st >= 500:
            if ps_index < len(ps_chain) - 1:
                ps_index += 1
                print(f"      ⚠️ st={st} — reduzindo pageSize para {ps_chain[ps_index]} e repetindo pág {page}…")
                continue
            print("      ⚠️ Falha persistente — tentando sem $filter e filtrando localmente…")
            items2, st2, _ = _get_movements_page(page, page_size, None, select_fields)
            if st2 >= 400 or st2 == -1:
                print(f"      ⚠️ HTTP {st2} mesmo sem filtro — encerrando filial.")
                break
            items = items2
            # filtros locais
            if comp_field:
                items = [it for it in items if str(it.get(comp_field)) == _norm_code(comp_norm)]
            if br_field and str(b_code).strip() != "":
                items = [it for it in items if str(it.get(br_field, "")).strip() == str(b_code).strip()]
            if not used_date_field:
                items = [it for it in items if in_period_local(it)]
        elif st >= 400:
            print(f"      ⚠️ HTTP {st} — pulando filial.")
            break

        if not items:
            empty_seq += 1
            print(f"    · Página {page}: 0 itens (vazias seguidas={empty_seq}) | ps={page_size}")
            if empty_seq >= 2:
                print("      ↳ Duas páginas vazias — fim desta filial.")
                break
            page += 1
            continue

        if not used_date_field:
            items = [it for it in items if in_period_local(it)]

        # dedupe por filial
        deduped = []
        for it in items:
            rid = None
            for f in id_fields:
                if f in it: rid = str(it[f]); break
            if rid and rid in seen_ids:
                continue
            if rid: seen_ids.add(rid)
            deduped.append(it)
        items = deduped

        out.extend(items)
        print(f"    · Página {page}: {len(items)} itens (acum={len(out)}) | ps={page_size}")
        if global_limit and len(out) >= global_limit:
            print(f"      ⏹️ Limite global de {global_limit} registros atingido — encerrando coleta desta filial.")
            out = out[:global_limit]
            break


        if len(items) < page_size:
            print("      ↳ Última página (len(items) < pageSize).")
            break
        page += 1

    return out

# ==============================
# MAIN
# ==============================
def run_extract(args):
    TZBR = tz.gettz("America/Sao_Paulo")
    dt_ini = datetime.strptime(args.inicio, "%Y-%m-%d").replace(hour=0, minute=0, second=0, tzinfo=TZBR)
    dt_fim = datetime.strptime(args.fim, "%Y-%m-%d").replace(hour=23, minute=59, second=59, tzinfo=TZBR)

    print(">>> Iniciando extração de movements…")
    print(f"Período: {dt_ini.isoformat()}  até  {dt_fim.isoformat()}")
    print(f"Endpoint: {HOST}{MOV_ENDPOINT}")

    # empresas alvo
    if args.empresa:
        target_companies = [_norm_code(x) for x in args.empresa.split(",")]

        print(f"🏭 Empresa informada por parâmetro: {target_companies[0]}")
        companies_meta = []
    else:
        companies = list_companies()
        if not companies:
            target_companies = ["1", "148"]
            print("⚠️ Não consegui listar empresas automaticamente; usando fallback ['4','148'].")
            companies_meta = []
        else:
            target_companies = [c["Code"] for c in companies]
            companies_meta = companies
            print(f"🏭 Empresas detectadas automaticamente: {', '.join(target_companies)}")

    # $select
    if args.select_fields == "":
        select_fields = []
    elif args.select_fields is None:
        select_fields = DEFAULT_SELECT_FIELDS[:]
    else:
        select_fields = [c.strip() for c in args.select_fields.split(",") if c.strip()]

    page_size_first = max(1, args.page_size)
    branch_deadline = max(0, args.branch_deadline_seconds)

    all_rows: List[Dict[str,Any]] = []
    all_rows_lock = Lock()

    # função por empresa (filiais em paralelo; empresas sequenciais → logs mais legíveis)
    def process_company(comp):
        comp_norm = _norm_code(comp)
        comp_name = get_company_name(comp_norm)
        if not comp_name and companies_meta:
            try:
                comp_name = next((c["Name"] for c in companies_meta if c["Code"] == comp_norm), "")
            except Exception:
                pass

        print("\n" + "="*72)
        print(f"🏢 Empresa {comp_norm} — {comp_name or '(sem nome)'}")
        print("="*72)

        branches = list_branches(comp_norm)
        print(f"  📚 Filiais detectadas: {len(branches)}")

        if args.filial:
            branches = [b for b in branches if str(b.get("Code","")).strip() == str(args.filial).strip()]
            print(f"  🔎 Filtro de filial: {args.filial} → {len(branches)} filial(is)")
        if not branches:
            print("  ⚠️ Nenhuma filial após filtro. Pulando empresa.")
            return

        # probe /movements para detectar campos
        probe_items, status, _ = _get_movements_page(1, 5, None, select_fields)
        if status >= 400 or status == -1 or not probe_items:
            probe_items2, status2, _ = _get_movements_page(1, 1, None, select_fields)
            if status2 >= 400 or status2 == -1 or not probe_items2:
                print(f"  ❌ Probe /movements falhou (HTTP {status2}).")
                return
            probe_items = probe_items2

        probe_row = probe_items[0] if probe_items else {}
        comp_field = _detect_field_present(probe_row, COMPANY_KEYS) or "companyId"
        br_field   = _detect_field_present(probe_row, BRANCH_KEYS)  or "branchId"
        date_candidates = _guess_datetime_fields_from_sample(probe_items)
        print(f"  🔎 Campos /movements → company='{comp_field}', branch='{br_field}'")

        # filiais em paralelo
        futures = []
        with ThreadPoolExecutor(max_workers=args.concurrency_branches) as ex:
            for b in branches:
                b_code = b.get("Code","")
                futures.append(ex.submit(
                    extract_branch_items,
                    comp_norm, b_code, comp_field, br_field,
                    dt_ini, dt_fim, date_candidates,
                    page_size_first, select_fields, branch_deadline
                ))

            total_emp = 0
            for fut in as_completed(futures):
                try:
                    items = fut.result()
                except Exception as e:
                    print(f"    ⚠️ Erro processando filial: {e}")
                    continue
                with all_rows_lock:
                    all_rows.extend(items)
                    total_emp += len(items)
                print(f"  ✅ Parcial empresa {comp_norm}: +{len(items)} itens (acum empresa={total_emp})")

        print(f"  ✅ Empresa {comp_norm} finalizada.")

    # percorre TODAS as empresas (sequencialmente para logs limpos)
    for comp in target_companies:
        process_company(comp)

    # ==============================
    # CONSOLIDAÇÃO
    # ==============================
    print("\n>>> Consolidando em DataFrame…")
    if not all_rows:
        print("(Sem registros coletados.)")
        return

    df = json_normalize(all_rows, sep=".")
    # ============================================
    # 🔍 Enriquecimento: busca CNPJ e Razão Social
    # ============================================
    print("🔍 Enriquecendo com CNPJ e Razão Social...")

    if "customerVendorCode" in df.columns and "companyId" in df.columns:
        unique_pairs = df[["customerVendorCode", "companyId"]].dropna().drop_duplicates()
        lookup = {}

        for _, row in unique_pairs.iterrows():
            code = str(row["customerVendorCode"]).strip()
            comp = str(row["companyId"]).strip()
            info = get_customer_vendor_info(code, comp)
            if info:
                lookup[(code, comp)] = {
                    "cnpj": info.get("mainNIF"),
                    "razao_social": info.get("shortName") or info.get("name")
                }

        df["cnpj"] = df.apply(lambda r: lookup.get((str(r["customerVendorCode"]), str(r["companyId"])), {}).get("cnpj"), axis=1)
        df["razao_social"] = df.apply(lambda r: lookup.get((str(r["customerVendorCode"]), str(r["companyId"])), {}).get("razao_social"), axis=1)
    else:
        print("⚠️ Campos customerVendorCode ou companyId não encontrados; não foi possível enriquecer.")

    df = _coerce_datetime_cols(df)
    df = _strip_timezone_for_excel(df)

    ordenar_cols = [c.strip() for c in (args.ordenar or "").split(",") if c.strip()]
    df = _smart_sort(df, ordenar_cols)

    branch_col  = _first_existing(["branchId","BranchId","branchCode","BranchCode"], set(df.columns))
    company_col = _first_existing(["companyId","CompanyId","companyCode","CompanyCode"], set(df.columns))
    resumo = None
    if branch_col:
        grp_cols = [c for c in [company_col, branch_col] if c]
        resumo = df.groupby(grp_cols, dropna=False).size().reset_index(name="qtd").sort_values("qtd", ascending=False)
        resumo = _strip_timezone_for_excel(resumo)

    # ==============================
    # XLSX
    # ==============================
    base = args.arquivo
    if not base:
        emp = f"emp{args.empresa}" if args.empresa else "empVARIAS"
        fil = f"_fil{args.filial}" if args.filial else ""
        base = f"movements_{emp}{fil}_{args.inicio}_a_{args.fim}.xlsx"
    if not base.lower().endswith(".xlsx"):
        base += ".xlsx"

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws_tree = wb.active
    ws_tree.title = "Árvore"
    ws_data = wb.create_sheet("Dados")

    df_excel = _jsonify_non_scalars_for_excel(df.copy())
    ws_data.append(list(df_excel.columns))
    for row in df_excel.itertuples(index=False, name=None):
        ws_data.append(list(row))
    try: _autosize_openpyxl(ws_data)
    except Exception: pass

    bold = Font(bold=True)
    header_fill = PatternFill(start_color="FFECECEC", end_color="FFECECEC", fill_type="solid")
    company_fill = PatternFill(start_color="FFCCE5FF", end_color="FFCCE5FF", fill_type="solid")
    branch_fill  = PatternFill(start_color="FFFFF2CC", end_color="FFFFF2CC", fill_type="solid")

    tree_cols = [c.strip() for c in (args.colunas or "").split(",") if c.strip()]
    if not tree_cols:
        pref = ["date","internalId","movementId","documentNumber","type","grossValue","netValue","partner.name","status"]
        tree_cols = [c for c in pref if c in df.columns] or list(df.columns[:10])

    comp_series = df[company_col] if company_col in df.columns else pd.Series([""]*len(df))
    br_series   = df[branch_col]  if branch_col in df.columns else pd.Series([""]*len(df))
    unique_pairs = pd.DataFrame({"comp": comp_series, "branch": br_series}).drop_duplicates()

    filial_label_map: Dict[Tuple[str,str], str] = {}

    for comp_id in sorted(unique_pairs["comp"].astype(str).unique(), key=lambda x: (x is None, x)):
        comp_mask = (comp_series.astype(str) == str(comp_id))
        df_comp = df[comp_mask]

        ws_tree.append([f"Empresa: {comp_id}"])
        for cell in ws_tree[ws_tree.max_row]:
            cell.font = bold; cell.fill = company_fill

        filiais_comp = sorted(df_comp[branch_col].astype(str).unique()) if branch_col in df_comp.columns else [""]
        for br_id in filiais_comp:
            br_mask = (br_series.astype(str) == str(br_id))

            df_fil = df_comp[br_mask] if branch_col in df_comp.columns else df_comp

            br_name = filial_label_map.get((str(comp_id), str(br_id)), "")
            ws_tree.append([f"  Filial: {br_id} {('- ' + br_name) if br_name else ''}"])
            for cell in ws_tree[ws_tree.max_row]:
                cell.font = bold; cell.fill = branch_fill

            ws_tree.append([""] + tree_cols)
            for cell in ws_tree[ws_tree.max_row]:
                cell.font = bold; cell.fill = header_fill

            df_tree = df_fil[tree_cols].copy()
            df_tree = _jsonify_non_scalars_for_excel(df_tree)
            for _, rec in df_tree.iterrows():
                ws_tree.append([""] + [rec.get(c) for c in tree_cols])

            ws_tree.append([])

    try: _autosize_openpyxl(ws_tree)
    except Exception: pass

    if resumo is not None:
        ws_resumo = wb.create_sheet("Resumo")
        resumo_excel = _jsonify_non_scalars_for_excel(resumo.copy())
        ws_resumo.append(list(resumo_excel.columns))
        for row in resumo_excel.itertuples(index=False, name=None):
            ws_resumo.append(list(row))
        try: _autosize_openpyxl(ws_resumo)
        except Exception: pass

    try:
        # 1. cria arquivo temporário
        tmpfile = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx").name

        # 2. salva no temp
        wb.save(tmpfile)
        wb.close()  # fecha handle do openpyxl

        # 3. move para destino final (substitui se existir)
        shutil.move(tmpfile, base)

    except Exception as e:
        alt = os.path.splitext(base)[0] + ".csv"
        print(f"⚠️ Falha ao salvar XLSX ({e}). Salvando CSV com dados completos: {alt}")
        df_excel.to_csv(alt, index=False)
        base = alt

    print(f"\n✅ Arquivo gerado: {os.path.abspath(base)}")

def parse_args():
    ap = argparse.ArgumentParser(description="Extrai Movements e gera XLSX em árvore + dados completos (rápido & estável).")
    ap.add_argument("--inicio",   required=True, help="Data inicial (YYYY-MM-DD).")
    ap.add_argument("--fim",      required=True, help="Data final (YYYY-MM-DD).")
    ap.add_argument("--empresa", required=False, help="IDs de empresas separados por vírgula, ex.: 4,147,148,149")

    ap.add_argument("--filial",   required=False, help="branchId (filial), ex.: 17")
    ap.add_argument("--ordenar",  required=False, default="date,internalId", help="Colunas para ordenar (vírgula).")
    ap.add_argument("--colunas",  required=False, default="", help="Colunas para mostrar na aba Árvore (vírgula).")
    ap.add_argument("--arquivo",  required=False, default="", help="Caminho do .xlsx de saída.")

    # performance/completude
    ap.add_argument("--page-size", type=int, default=100, help="Itens por página (o servidor pode limitar; varre até o fim).")
    ap.add_argument("--concurrency-branches", type=int, default=4, help="Filiais processadas em paralelo.")
    ap.add_argument("--branch-deadline-seconds", type=int, default=BRANCH_DEADLINE_SEC_DEFAULT,
                    help="Tempo máx por filial; 0 = sem limite (não perder itens).")

    # $select (string vazia desliga)
    ap.add_argument("--select-fields", type=str, default=",".join(DEFAULT_SELECT_FIELDS),
                    help='Campos em $select (ex.: "date,internalId,movementId,companyId,branchId"). Use "" para desativar.')
    ap.add_argument("--movimento", required=False, help="Lista de códigos de movementTypeCode separados por vírgula, ex.: 1.2.01,1.2.09,1.2.10.")
    ap.add_argument("--limit", type=int, default=0, help="Limite máximo de registros a coletar (0 = sem limite).")
    ap.add_argument("--max-pages", type=int, default=1, help="Número máximo de páginas a buscar por filial (padrão=1).")



    return ap.parse_args()

if __name__ == "__main__":
    args = parse_args()
    try:
        run_extract(args)
    except KeyboardInterrupt:
        print("\nInterrompido pelo usuário.")
        sys.exit(130)
    except Exception as e:
        print(f"\nERRO fatal: {e}")
        sys.exit(2)
